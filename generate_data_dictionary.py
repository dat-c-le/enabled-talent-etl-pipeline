"""
Generate a formatted Excel data dictionary for the disability employment ETL pipeline.
Output: output/Disability_Employment_Data_Dictionary.xlsx
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path("output") / "Disability_Employment_Data_Dictionary.xlsx"

# ── Colour palette ─────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1F4E79"   # dark navy  — column header rows
CLR_HEADER_FG   = "FFFFFF"
CLR_SECTION_BG  = "2E75B6"   # mid blue   — section sub-headers
CLR_SECTION_FG  = "FFFFFF"
CLR_TABLE_BG    = "D6E4F0"   # light blue — table title bar
CLR_ALT_ROW     = "EBF5FB"   # very light blue — alternating data rows
CLR_NOTE_BG     = "FFF2CC"   # yellow     — year-availability note rows
CLR_NOTE_FG     = "7F6000"

# ── Thin border helper ──────────────────────────────────────────────────────────
_thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def hdr_font(bold=True, size=11, color=CLR_HEADER_FG):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def body_font(bold=False, size=10, italic=False, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, italic=italic, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def wrap_align(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)


def style_cell(cell, value, font=None, bg=None, alignment=None, border=True):
    cell.value = value
    if font:
        cell.font = font
    if bg:
        cell.fill = fill(bg)
    cell.alignment = alignment or wrap_align()
    if border:
        cell.border = BORDER


def write_sheet(wb, sheet_name, table_id, description, universe, years, notes, sections):
    """
    sections: list of (section_title, rows)
      rows: list of (column_name, data_type, description, year_note)
             year_note may be "" or a short string like "2015–2024 only"
    """
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36   # column name
    ws.column_dimensions["B"].width = 14   # data type
    ws.column_dimensions["C"].width = 62   # description
    ws.column_dimensions["D"].width = 22   # year availability

    row = 1

    # ── Table title banner ─────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1)
    style_cell(c, f"{table_id}  —  {description}",
               font=Font(name="Calibri", bold=True, size=13, color="FFFFFF"),
               bg=CLR_HEADER_BG,
               alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
               border=False)
    ws.row_dimensions[row].height = 24
    row += 1

    # ── Metadata rows ──────────────────────────────────────────────────────────
    meta = [("Universe", universe), ("Years covered", years)]
    for label, value in meta:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1)
        style_cell(c, f"  {label}:  {value}",
                   font=body_font(bold=False, size=10, color="1F4E79"),
                   bg="D6E4F0",
                   alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                   border=False)
        ws.row_dimensions[row].height = 16
        row += 1

    # ── Year-coverage note (if any) ────────────────────────────────────────────
    if notes:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1)
        style_cell(c, f"  Note:  {notes}",
                   font=body_font(bold=False, size=9, italic=True, color=CLR_NOTE_FG),
                   bg=CLR_NOTE_BG,
                   alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                   border=False)
        ws.row_dimensions[row].height = 40
        row += 1

    row += 1   # spacer

    # ── Column header row ──────────────────────────────────────────────────────
    headers = ["Column Name", "Data Type", "Description", "Year Availability"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx)
        style_cell(c, h,
                   font=hdr_font(size=10),
                   bg=CLR_HEADER_BG,
                   alignment=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[row].height = 18
    row += 1

    # ── Sections and data rows ─────────────────────────────────────────────────
    alt = False
    for section_title, rows_data in sections:
        # Section sub-header
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1)
        style_cell(c, f"  {section_title}",
                   font=Font(name="Calibri", bold=True, size=10, color=CLR_SECTION_FG),
                   bg=CLR_SECTION_BG,
                   alignment=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[row].height = 16
        row += 1

        for col_name, dtype, desc, year_note in rows_data:
            bg = CLR_ALT_ROW if alt else "FFFFFF"
            vals = [col_name, dtype, desc, year_note]
            for col_idx, val in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col_idx)
                bold = col_idx == 1
                style_cell(c, val,
                           font=body_font(bold=bold, size=10),
                           bg=bg,
                           alignment=wrap_align())
            ws.row_dimensions[row].height = 30
            row += 1
            alt = not alt

    # ── Freeze panes below header row ─────────────────────────────────────────
    # Find the header row index (it's row 4 or 5 depending on notes)
    # Just freeze at row 6 (safe)
    ws.freeze_panes = f"A{7 if notes else 6}"

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# Data definitions
# ══════════════════════════════════════════════════════════════════════════════

SHARED_COLUMNS = [
    ("year",         "integer", "ACS 1-year estimate vintage (2010–2024, no 2020)",                              "All years"),
    ("survey_type",  "string",  "Always '1-Year'",                                                              "All years"),
    ("geo_id",       "string",  "Census GEOID in UCGID format (e.g. 0400000US01)",                              "All years"),
    ("level",        "string",  "'state' or 'county'",                                                          "All years"),
    ("state",        "string",  "State name (e.g. 'Alabama')",                                                  "All years"),
    ("state_fips",   "integer", "2-digit state FIPS code",                                                      "All years"),
    ("county",       "string",  "County name — null for state-level rows",                                      "All years"),
    ("county_fips",  "integer", "3-digit county FIPS — null for state-level rows",                              "All years"),
    ("fips",         "string",  "Full FIPS: state FIPS for states, state+county FIPS for counties",             "All years"),
]

S1810_SECTIONS = [
    ("Geographic & Identifier Columns", [
        (n, t, d, y) for n, t, d, y in SHARED_COLUMNS
    ]),
    ("Total Population", [
        ("pop_total",       "integer", "Total civilian noninstitutionalized population",    "All years"),
        ("pop_male",        "integer", "Total population — male",                           "All years"),
        ("pop_female",      "integer", "Total population — female",                         "All years"),
        ("pop_age_18_34",   "integer", "Total population age 18–34",                        "2015–2024 only"),
        ("pop_age_35_64",   "integer", "Total population age 35–64",                        "2015–2024 only"),
        ("pop_age_18_64",   "integer", "Total population age 18–64",                        "2010–2014 only"),
        ("pop_age_65_plus", "integer", "Total population age 65 and over",                  "2010–2014 only"),
    ]),
    ("Population with a Disability", [
        ("dis_pop_total",        "integer", "Population with any disability",                     "All years"),
        ("dis_pop_male",         "integer", "Population with disability — male",                   "All years"),
        ("dis_pop_female",       "integer", "Population with disability — female",                  "All years"),
        ("dis_pop_white",        "integer", "Population with disability — White alone",             "All years"),
        ("dis_pop_black",        "integer", "Population with disability — Black or African American alone", "All years"),
        ("dis_pop_asian",        "integer", "Population with disability — Asian alone",             "All years"),
        ("dis_pop_hispanic",     "integer", "Population with disability — Hispanic or Latino (any race)", "All years"),
        ("dis_pop_age_18_34",    "integer", "Population with disability age 18–34",                 "2015–2024 only"),
        ("dis_pop_age_35_64",    "integer", "Population with disability age 35–64",                 "2015–2024 only"),
        ("dis_pop_age_65_74",    "integer", "Population with disability age 65–74",                 "2015–2024 only"),
        ("dis_pop_age_18_64",    "integer", "Population with disability age 18–64",                 "2010–2014 only"),
        ("dis_pop_age_65_plus",  "integer", "Population with disability age 65 and over",           "2010–2014 only"),
    ]),
    ("Disability Type (2015–2024 only)", [
        ("dis_type_hearing_total",       "integer", "Population with a hearing difficulty — all ages",          "2015–2024 only"),
        ("dis_type_hearing_18_64",       "integer", "Population with a hearing difficulty — age 18–64",         "2015–2024 only"),
        ("dis_type_vision_total",        "integer", "Population with a vision difficulty — all ages",           "2015–2024 only"),
        ("dis_type_vision_18_64",        "integer", "Population with a vision difficulty — age 18–64",          "2015–2024 only"),
        ("dis_type_cognitive_total",     "integer", "Population with a cognitive difficulty — all ages",        "2015–2024 only"),
        ("dis_type_cognitive_18_64",     "integer", "Population with a cognitive difficulty — age 18–64",       "2015–2024 only"),
        ("dis_type_ambulatory_total",    "integer", "Population with an ambulatory (mobility) difficulty — all ages", "2015–2024 only"),
        ("dis_type_ambulatory_18_64",    "integer", "Population with an ambulatory difficulty — age 18–64",     "2015–2024 only"),
        ("dis_type_self_care_total",     "integer", "Population with a self-care difficulty — all ages",        "2015–2024 only"),
        ("dis_type_self_care_18_64",     "integer", "Population with a self-care difficulty — age 18–64",       "2015–2024 only"),
        ("dis_type_indep_living_18_64",  "integer", "Population with an independent living difficulty — age 18–64", "2015–2024 only"),
    ]),
]

S1811_SECTIONS = [
    ("Geographic & Identifier Columns", [
        (n, t, d, y) for n, t, d, y in SHARED_COLUMNS
    ]),
    ("Population Totals", [
        ("pop_16_plus",          "integer", "Total civilian population age 16 and over",                  "All years"),
        ("pop_employed",         "integer", "Total employed population age 16 and over",                  "All years"),
        ("dis_pop_16_plus",      "integer", "Population with disability age 16 and over",                 "All years"),
        ("dis_employed_total",   "integer", "Employed population with disability age 16 and over",        "All years"),
        ("nodis_pop_16_plus",    "integer", "Population without disability age 16 and over",              "All years"),
        ("nodis_employed_total", "integer", "Employed population without disability age 16 and over",     "All years"),
    ]),
    ("Employment Status", [
        ("dis_employed",     "integer", "Population with disability — employed",          "All years"),
        ("dis_not_in_lf",    "integer", "Population with disability — not in labor force", "All years"),
        ("nodis_employed",   "integer", "Population without disability — employed",        "All years"),
    ]),
    ("Class of Worker", [
        ("dis_sector_private_forprofit", "integer", "Employed with disability — private for-profit wage & salary workers", "All years"),
        ("dis_sector_private_employee",  "integer", "Employed with disability — employee of private company",              "2013–2024 only"),
        ("dis_sector_self_emp_inc",      "integer", "Employed with disability — self-employed, incorporated business",     "2013–2024 only"),
        ("dis_sector_nonprofit",         "integer", "Employed with disability — private not-for-profit wage & salary workers", "All years"),
        ("dis_sector_local_govt",        "integer", "Employed with disability — local government workers",                 "All years"),
        ("dis_sector_state_govt",        "integer", "Employed with disability — state government workers",                 "All years"),
        ("dis_sector_federal_govt",      "integer", "Employed with disability — federal government workers",               "All years"),
        ("dis_sector_self_emp_uninc",    "integer", "Employed with disability — self-employed, unincorporated business",   "All years"),
        ("dis_sector_unpaid_family",     "integer", "Employed with disability — unpaid family workers",                    "All years"),
    ]),
    ("Occupation", [
        ("dis_occ_management",        "integer", "Employed with disability — management, business, science & arts",                   "All years"),
        ("dis_occ_service",           "integer", "Employed with disability — service occupations",                                    "All years"),
        ("dis_occ_sales_office",      "integer", "Employed with disability — sales and office occupations",                           "All years"),
        ("dis_occ_natural_resources", "integer", "Employed with disability — natural resources, construction & maintenance",          "All years"),
        ("dis_occ_production",        "integer", "Employed with disability — production, transportation & material moving",           "All years"),
    ]),
    ("Industry", [
        ("dis_ind_agriculture",      "integer", "Employed with disability — agriculture, forestry, fishing, hunting & mining", "All years"),
        ("dis_ind_construction",     "integer", "Employed with disability — construction",                                     "All years"),
        ("dis_ind_manufacturing",    "integer", "Employed with disability — manufacturing",                                    "All years"),
        ("dis_ind_wholesale",        "integer", "Employed with disability — wholesale trade",                                  "All years"),
        ("dis_ind_retail",           "integer", "Employed with disability — retail trade",                                    "All years"),
        ("dis_ind_transportation",   "integer", "Employed with disability — transportation, warehousing & utilities",          "All years"),
        ("dis_ind_information",      "integer", "Employed with disability — information",                                     "All years"),
        ("dis_ind_finance",          "integer", "Employed with disability — finance, insurance, real estate & rental",         "All years"),
        ("dis_ind_professional",     "integer", "Employed with disability — professional, scientific, management & administrative", "All years"),
        ("dis_ind_education_health", "integer", "Employed with disability — educational services, health care & social assistance", "All years"),
        ("dis_ind_arts_food",        "integer", "Employed with disability — arts, entertainment, recreation & food services",  "All years"),
        ("dis_ind_other_services",   "integer", "Employed with disability — other services (except public administration)",    "All years"),
        ("dis_ind_public_admin",     "integer", "Employed with disability — public administration",                           "All years"),
    ]),
    ("Educational Attainment (2013–2024 only)", [
        ("dis_edu_pop_25_plus",      "integer", "Population with disability age 25 and over (education universe base)", "2013–2024 only"),
        ("dis_edu_less_than_hs",     "integer", "Population with disability — less than high school graduate",          "2013–2024 only"),
        ("dis_edu_hs_grad",          "integer", "Population with disability — high school graduate or equivalency",     "2013–2024 only"),
        ("dis_edu_some_college",     "integer", "Population with disability — some college or associate's degree",      "2013–2024 only"),
        ("dis_edu_bachelors_plus",   "integer", "Population with disability — bachelor's degree or higher",             "2013–2024 only"),
        ("nodis_edu_bachelors_plus", "integer", "Population without disability — bachelor's degree or higher",          "2013–2024 only"),
    ]),
    ("Commuting (2019–2024 only)", [
        ("dis_work_from_home",   "integer", "Employed with disability — worked from home",    "2019–2024 only"),
        ("nodis_work_from_home", "integer", "Employed without disability — worked from home", "2019–2024 only"),
    ]),
]

B18120_SECTIONS = [
    ("Geographic & Identifier Columns", [
        (n, t, d, y) for n, t, d, y in SHARED_COLUMNS
    ]),
    ("Total Population", [
        ("pop_total",          "integer", "Total civilian noninstitutionalized population age 18–64", "All years"),
        ("in_labor_force",     "integer", "In the labor force",                                       "All years"),
        ("employed_total",     "integer", "In the labor force — employed",                            "All years"),
        ("not_in_labor_force", "integer", "Not in the labor force",                                   "All years"),
    ]),
    ("Employed", [
        ("dis_employed",               "integer", "Employed — with any disability",               "All years"),
        ("dis_employed_hearing",       "integer", "Employed — with a hearing difficulty",         "All years"),
        ("dis_employed_vision",        "integer", "Employed — with a vision difficulty",          "All years"),
        ("dis_employed_cognitive",     "integer", "Employed — with a cognitive difficulty",       "All years"),
        ("dis_employed_ambulatory",    "integer", "Employed — with an ambulatory difficulty",     "All years"),
        ("dis_employed_self_care",     "integer", "Employed — with a self-care difficulty",       "All years"),
        ("dis_employed_indep_living",  "integer", "Employed — with an independent living difficulty", "All years"),
        ("nodis_employed",             "integer", "Employed — no disability",                     "All years"),
    ]),
    ("Unemployed & Not in Labor Force", [
        ("dis_unemployed",    "integer", "Unemployed — with any disability",                 "All years"),
        ("nodis_unemployed",  "integer", "Unemployed — no disability",                       "All years"),
        ("dis_not_in_lf",     "integer", "Not in the labor force — with any disability",     "All years"),
        ("nodis_not_in_lf",   "integer", "Not in the labor force — no disability",           "All years"),
    ]),
]

B18121_SECTIONS = [
    ("Geographic & Identifier Columns", [
        (n, t, d, y) for n, t, d, y in SHARED_COLUMNS
    ]),
    ("Total Population", [
        ("pop_total",     "integer", "Total civilian noninstitutionalized population age 18–64", "All years"),
        ("fulltime_total","integer", "Worked full-time, year round",                             "All years"),
        ("parttime_total","integer", "Worked less than full-time, year round",                   "All years"),
        ("did_not_work_total","integer","Did not work during the year",                          "All years"),
    ]),
    ("Full-Time, Year Round", [
        ("dis_fulltime",               "integer", "Worked full-time, year round — with any disability",               "All years"),
        ("dis_fulltime_hearing",       "integer", "Worked full-time, year round — with a hearing difficulty",         "All years"),
        ("dis_fulltime_vision",        "integer", "Worked full-time, year round — with a vision difficulty",          "All years"),
        ("dis_fulltime_cognitive",     "integer", "Worked full-time, year round — with a cognitive difficulty",       "All years"),
        ("dis_fulltime_ambulatory",    "integer", "Worked full-time, year round — with an ambulatory difficulty",     "All years"),
        ("dis_fulltime_self_care",     "integer", "Worked full-time, year round — with a self-care difficulty",       "All years"),
        ("dis_fulltime_indep_living",  "integer", "Worked full-time, year round — with an independent living difficulty", "All years"),
        ("nodis_fulltime",             "integer", "Worked full-time, year round — no disability",                     "All years"),
    ]),
    ("Less Than Full-Time & Did Not Work", [
        ("dis_parttime",       "integer", "Worked less than full-time, year round — with any disability",  "All years"),
        ("nodis_parttime",     "integer", "Worked less than full-time, year round — no disability",        "All years"),
        ("dis_did_not_work",   "integer", "Did not work — with any disability",                            "All years"),
        ("nodis_did_not_work", "integer", "Did not work — no disability",                                  "All years"),
    ]),
]

CPS_SECTIONS = [
    ("Series Identification", [
        ("series_id",     "string", "BLS LABSTAT series ID. 'LNU0' prefix = not seasonally adjusted; 'LNS1' prefix = seasonally adjusted", "All years"),
        ("series_title",  "string", "Full BLS series name as published",                                                                 "All years"),
        ("periodicity",   "string", "Series frequency: 'M' monthly, 'A' annual, 'Q' quarterly",                                            "All years"),
        ("seasonal_adjustment", "string", "'Seasonally adjusted' or 'Not seasonally adjusted'. Disability-specific series are NSA only — BLS does not publish SA disability series (sample too small)", "All years"),
    ]),
    ("Demographic Dimensions", [
        ("disability_status", "string", "'With disability' / 'No disability' / 'All persons' (general population total)", "All years"),
        ("sex",            "string", "'Both sexes' / 'Men' / 'Women'",                                                    "All years"),
        ("age_group",       "string", "Age range of the population measured, e.g. '16 years and over', '25 to 34 years'", "All years"),
        ("race_ethnicity",  "string", "'All races/ethnicities' / 'White' / 'Black or African American' / 'Asian' / 'Hispanic or Latino'. Hispanic/Latino takes priority over race when both apply. Breakdown only available for disability totals (annual, 16+, both sexes)", "All years"),
    ]),
    ("Labor Force Measure", [
        ("labor_force_status", "string", "The labor market measure reported: Population, Civilian labor force, Employed, Employed full/part time, Unemployed, Not in labor force, Labor force participation rate, Employment-population ratio, Unemployment rate. Rate/ratio values are percent (0-100); all others are thousands of persons", "All years"),
    ]),
    ("Table 3/4/5 Breakdowns (annual, With/No disability only)", [
        ("occupation",       "string", "Occupation group, e.g. 'Management occupations', 'Service occupations' (Table 3). Empty for non-occupation series. Value converted from BLS percent distribution to a count in thousands using total employed as denominator", "Annual only"),
        ("industry",         "string", "Industry sector, e.g. 'Construction', 'Manufacturing' (Table 4, With disability only). Empty for non-industry series. Value converted to a count in thousands",                                                       "Annual only"),
        ("class_of_worker",  "string", "Class of worker, e.g. 'Wage and salary workers', 'Self-employed Workers, Unincorporated' (Table 4, With disability only). Value converted to a count in thousands",                                              "Annual only"),
        ("nilf_subcategory", "string", "Not-in-labor-force detail or education level, e.g. 'Discouraged workers', 'Bachelor's degree or higher' (Table 5, With disability only)",                                                                          "Annual only"),
    ]),
    ("Time & Value", [
        ("year",       "integer", "Calendar year",                                                          "All years"),
        ("period",     "string",  "BLS period code: 'M01'-'M12' (month), 'Q01'-'Q04' (quarter), 'A01' (annual)", "All years"),
        ("date",       "string",  "ISO date of the first day of the period (e.g. '2022-03-01' for March)",  "All years"),
        ("value",      "float",   "Measured value. Units depend on labor_force_status: thousands of persons for counts, percent (0-100) for rates/ratios. occupation/industry/class_of_worker values are always counts in thousands", "All years"),
        ("footnotes",  "string",  "BLS footnote codes, if any. Mostly null",                                 "All years"),
    ]),
]


# ══════════════════════════════════════════════════════════════════════════════
# Build workbook
# ══════════════════════════════════════════════════════════════════════════════

def build_overview_sheet(wb):
    ws = wb.create_sheet(title="Overview", index=0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 10

    row = 1

    # Title
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1)
    style_cell(c, "Disability Employment — ACS Data Dictionary",
               font=Font(name="Calibri", bold=True, size=14, color="FFFFFF"),
               bg=CLR_HEADER_BG,
               alignment=Alignment(horizontal="center", vertical="center"),
               border=False)
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1)
    style_cell(c, "  Sources: U.S. Census Bureau ACS 1-Year Estimates (state/county) and BLS Current Population Survey (national)",
               font=body_font(size=10, color="1F4E79"),
               bg="D6E4F0",
               alignment=Alignment(horizontal="left", vertical="center"),
               border=False)
    ws.row_dimensions[row].height = 16
    row += 2

    # Table of contents header
    for col_idx, h in enumerate(["Table ID", "Description", "Universe", "Years", "# Columns"], start=1):
        c = ws.cell(row=row, column=col_idx)
        style_cell(c, h,
                   font=hdr_font(size=10),
                   bg=CLR_HEADER_BG,
                   alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[row].height = 18
    row += 1

    tables = [
        ("S1810", "Disability Characteristics",
         "Total civilian noninstitutionalized population",
         "2010–2024", "30"),
        ("S1811", "Employment & Economic Characteristics by Disability Status",
         "Civilian noninstitutionalized population age 16+",
         "2010–2024", "44"),
        ("B18120", "Employment Status by Disability Status and Type",
         "Civilian noninstitutionalized population age 18–64",
         "2010–2024", "16"),
        ("B18121", "Work Experience by Disability Status and Type",
         "Civilian noninstitutionalized population age 18–64",
         "2010–2024", "16"),
        ("BLS CPS", "Disability Labor Force Characteristics (national)",
         "Civilian noninstitutional population age 16+ (national only)",
         "2008–2024", "18"),
    ]

    for i, (tid, desc, universe, years, ncols) in enumerate(tables):
        bg = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        for col_idx, val in enumerate([tid, desc, universe, years, ncols], start=1):
            c = ws.cell(row=row, column=col_idx)
            style_cell(c, val,
                       font=body_font(bold=(col_idx == 1), size=10),
                       bg=bg,
                       alignment=Alignment(horizontal="center" if col_idx in (1, 4, 5) else "left",
                                           vertical="center", wrap_text=True))
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # Shared columns note
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1)
    style_cell(c, "  The four ACS tables share the same 9 geographic/identifier columns (year, survey_type, geo_id, level, state, state_fips, county, county_fips, fips). "
                  "The BLS CPS table is national-level only (no state/county breakdown) and uses a tidy/long format — one row per series x year x period. "
                  "See each sheet for full column detail.",
               font=body_font(size=9, italic=True, color=CLR_NOTE_FG),
               bg=CLR_NOTE_BG,
               alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
               border=False)
    ws.row_dimensions[row].height = 40

    ws.freeze_panes = "A4"


def main():
    Path("output").mkdir(exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    build_overview_sheet(wb)

    write_sheet(
        wb, sheet_name="S1810",
        table_id="S1810",
        description="Disability Characteristics",
        universe="Total civilian noninstitutionalized population",
        years="2010–2024 (no 2020)",
        notes=(
            "Age sub-buckets differ by era: pop_age_18_64 / dis_pop_age_18_64 (2010–2014) vs "
            "pop_age_18_34 / pop_age_35_64 / dis_pop_age_18_34 / dis_pop_age_35_64 / dis_pop_age_65_74 (2015–2024). "
            "Disability-type columns (dis_type_*) are 2015–2024 only; Census restructured the table in 2015."
        ),
        sections=S1810_SECTIONS,
    )

    write_sheet(
        wb, sheet_name="S1811",
        table_id="S1811",
        description="Employment & Economic Characteristics by Disability Status",
        universe="Civilian noninstitutionalized population age 16 and over",
        years="2010–2024 (no 2020)",
        notes=(
            "Sector / occupation / industry values for 2010–2017 were published as percentages by Census; "
            "the pipeline converts them to counts using the relevant employed-population base row. "
            "Education columns: 2013–2024. Work-from-home columns: 2019–2024."
        ),
        sections=S1811_SECTIONS,
    )

    write_sheet(
        wb, sheet_name="B18120",
        table_id="B18120",
        description="Employment Status by Disability Status and Type",
        universe="Civilian noninstitutionalized population age 18–64",
        years="2010–2024 (no 2020)",
        notes="",
        sections=B18120_SECTIONS,
    )

    write_sheet(
        wb, sheet_name="B18121",
        table_id="B18121",
        description="Work Experience by Disability Status and Type",
        universe="Civilian noninstitutionalized population age 18–64",
        years="2010–2024 (no 2020)",
        notes="",
        sections=B18121_SECTIONS,
    )

    write_sheet(
        wb, sheet_name="BLS_CPS",
        table_id="bls_cps_disability",
        description="Disability Labor Force Characteristics (BLS CPS)",
        universe="Civilian noninstitutional population age 16 and over — national level only",
        years="2008–2024 (monthly + annual; no gap in 2020)",
        notes=(
            "Tidy/long format — one row per series x year x period (815 series, 34,898 rows). "
            "Disability-specific series are not seasonally adjusted (BLS does not publish SA disability "
            "series; sample too small). occupation/industry/class_of_worker/nilf_subcategory columns are "
            "populated only for the relevant Table 3/4/5 series and are annual-only; all other rows leave "
            "them blank. Percent-distribution series were converted to counts in thousands using the "
            "annual-average total employed as denominator."
        ),
        sections=CPS_SECTIONS,
    )

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
